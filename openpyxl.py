import openpyxl

# Carregar a planilha
workbook = openpyxl.load_workbook('caminho/para/sua_planilha.xlsx')

# Selecionar a folha ativa (ou uma folha específica pelo nome)
sheet = workbook.active  # ou workbook['Nome_da_Folha']

# Ler uma célula específica
valor = sheet['A1'].value
print(f"Valor da célula A1: {valor}")

# Ler um intervalo específico de células
print("Lendo intervalo de células A1 a C2:")
for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

# Ler todas as células
print("Lendo todas as células:")
for row in sheet.iter_rows(values_only=True):
    print(row)